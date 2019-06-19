#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 19 21:01:21 2019

@author: sheikh
"""

import openpyxl
from sklearn.linear_model import LogisticRegression
import pickle

wb = openpyxl.load_workbook('CodingChallengeData.xlsx')


train = wb['Development']


x_train = []
for rowOfCellObjects in train['B2':'AH'+str(train.max_row)]:
    row = []
    for cellObj in rowOfCellObjects:
        row.append(cellObj.value)
    x_train.append(row)
    
y_train = []
for rowOfCellObjects in train['AI2':'AI'+str(train.max_row)]:
        for cellObj in rowOfCellObjects:
            y_train.append(cellObj.value)
    
#keep it simple, stupid
model = LogisticRegression()
model.fit(x_train, y_train)

        
filename = 'finalized_model.sav'
pickle.dump(model, open(filename, 'wb'))
    
    
    

    
