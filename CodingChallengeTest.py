#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 19 22:54:31 2019

@author: sheikh
"""
import openpyxl
from sklearn.externals import joblib


wb = openpyxl.load_workbook('CodingChallengeData.xlsx')
test = wb['Validation']


loaded_model = joblib.load('finalized_model.sav')

x_test = []
for rowOfCellObjects in test['B2':'AH'+str(test.max_row)]:
    row = []
    for cellObj in rowOfCellObjects:
        row.append(cellObj.value)
    x_test.append(row)
    
y_test = loaded_model.predict_proba(x_test)
y_test = [i*1000 for i in y_test]

#not really pythonic, but who cares
j = 0
for rowNum in range(2, len(y_test)):
    test.cell(row=rowNum, column=35).value = int(y_test[rowNum - 2][0])
    j = j + 1
        
wb.save('CodingChallengeAnswer.xlsx')